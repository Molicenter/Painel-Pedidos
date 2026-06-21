import streamlit as st
import pandas as pd
import io
import datetime
import streamlit.components.v1 as components
from streamlit_gsheets import GSheetsConnection
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Gestão de Pedidos - Açougue Peças",
    page_icon="🥩",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# INICIALIZAÇÃO DE VARIÁVEIS DE SESSÃO
# ─────────────────────────────────────────────
if 'reset_counter_acpecas' not in st.session_state:
    st.session_state['reset_counter_acpecas'] = 0

if 'usuario_logado_acpecas' not in st.session_state:
    st.session_state['usuario_logado_acpecas'] = None

# ─────────────────────────────────────────────
# CSS GLOBAL E DE IMPRESSÃO (PALETA VERMELHA)
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500;700&display=swap');

:root {
    --bg-main:        #0d1117;
    --bg-card:        #161b22;
    --bg-sidebar:     #0d1117;
    --red-dark:       #3a1010;
    --red-mid:        #6b1d1d;
    --red-accent:     #b32d2d;
    --red-bright:     #e63939;
    --red-glow:       rgba(179, 45, 45, .25);
    --text-primary:   #e6edf3;
    --text-muted:     #7d8590;
    --text-header:    #ffcccc;
    --border:         #21262d;
    --border-active:  #b32d2d;
    --row-hover:      rgba(179, 45, 45, .08);
    --row-selected:   rgba(179, 45, 45, .18);
}

.stApp, .main { background-color: var(--bg-main) !important; color: var(--text-primary) !important; }
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif !important; }
section[data-testid="stSidebar"] { background-color: var(--bg-sidebar) !important; border-right: 1px solid var(--border); }
section[data-testid="stSidebar"] * { color: var(--text-primary) !important; }
section[data-testid="stSidebar"] .stRadio label { font-size: 14px; }

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, var(--red-mid) 0%, var(--red-accent) 100%) !important;
    color: #fff !important;
    border: 1px solid var(--red-accent) !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    letter-spacing: .3px;
    transition: all .2s ease !important;
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 18px var(--red-glow) !important;
}
.stButton > button {
    background: var(--bg-card) !important;
    color: var(--text-primary) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    transition: all .2s ease !important;
}
.stButton > button:hover {
    border-color: var(--red-accent) !important;
    color: var(--red-bright) !important;
    transform: translateY(-1px) !important;
}
.stTextInput input, .stSelectbox > div > div {
    background-color: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    color: var(--text-primary) !important;
}
.stTextInput input:focus, .stSelectbox > div > div:focus-within {
    border-color: var(--red-accent) !important;
    box-shadow: 0 0 0 3px var(--red-glow) !important;
}
.title-input input {
    font-weight: 700 !important;
    font-size: 16px !important;
    color: var(--red-bright) !important;
    padding: 2px 8px !important;
    background: transparent !important;
    border: 1px dashed #21262d !important;
}
.title-input input:focus { border: 1px dashed var(--red-accent) !important; }

[data-testid="stDataEditor"] [data-testid="glideDataEditor"] .gdg-header-cell,
[data-testid="stDataEditor"] .dvn-stack .gdg-header {
    background-color: var(--red-dark) !important;
    color: var(--text-header) !important;
}

[data-testid="stDataEditor"] {
    border-radius: 10px !important;
    overflow: hidden;
    border: 1px solid var(--red-mid) !important;
    box-shadow: 0 4px 20px rgba(0,0,0,.4);
    font-size: 12px !important;
}
[data-testid="stDataEditor"] .gdg-cell.gdg-selected,
[data-testid="stDataEditor"] .gdg-cell[data-state="focused"],
[data-testid="stDataEditor"] .gdg-cell[aria-selected="true"] {
    background-color: var(--row-selected) !important;
    outline: 2px solid var(--red-accent) !important;
    outline-offset: -2px;
}
[data-testid="stDataEditor"] .gdg-row:hover .gdg-cell { background-color: var(--row-hover) !important; }

div[data-testid="stVerticalBlockBorderWrapper"] {
    background-color: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
    transition: box-shadow .25s ease, border-color .25s ease;
}
div[data-testid="stVerticalBlockBorderWrapper"]:hover {
    border-color: var(--red-mid) !important;
    box-shadow: 0 6px 24px rgba(0,0,0,.35) !important;
}
[data-testid="stMetric"] {
    background-color: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 10px 10px;
}
[data-testid="stMetricValue"] { color: var(--red-bright) !important; font-weight: 700; font-size: 1.8rem !important; }
[data-testid="stMetricLabel"] { color: var(--text-muted) !important; }

.topbar-loja {
    background: linear-gradient(90deg, var(--red-dark) 0%, #1a0808 100%);
    border: 1px solid var(--red-mid);
    border-radius: 10px;
    padding: 10px 18px;
    margin-bottom: 18px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.topbar-left { display: flex; align-items: center; gap: 12px; }
.topbar-title { font-size: 18px; font-weight: 700; color: var(--text-header); }
.topbar-sub { font-size: 11px; color: var(--text-muted); margin-top: 2px; }

/* REGRAS DE IMPRESSÃO ABSOLUTAS - COMPACTADAS PARA CABER EM 1 PÁGINA */
@media print {
    @page { margin: 4mm 8mm; }
    .stApp, .main, body, html {
        background-color: #ffffff !important;
        background-image: none !important;
        color: #000000 !important;
        padding: 0 !important;
        margin: 0 !important;
    }
    header, [data-testid="stSidebar"], [data-testid="stHeader"] { 
        display: none !important; 
    }
    
    /* DESLIGA TUDO O QUE FOR DO STREAMLIT NA TELA */
    [data-testid="stElementContainer"],
    [data-testid="stHorizontalBlock"],
    div[data-testid="stVerticalBlockBorderWrapper"] {
        display: none !important;
    }
    
    /* RELIGA APENAS A CAIXA QUE CONTÉM O PRINT-SECTION */
    [data-testid="stElementContainer"]:has(#print-section) {
        display: block !important;
        width: 100% !important;
    }
    
    /* FORMATAÇÃO DO PAPEL */
    #print-section {
        display: block !important;
        width: 100% !important;
    }
    #print-section h2 {
        font-size: 14px !important;
        margin: 0 0 5px 0 !important;
        padding-bottom: 3px !important;
        border-bottom: 1px solid #000 !important;
        color: #000 !important;
        display: block !important;
    }
    #print-section h3 {
        font-size: 12px !important;
        border-bottom: none !important;
        margin-top: 10px !important;
        margin-bottom: 3px !important;
        color: #000 !important;
    }
    .print-container { width: 100%; display: block !important;}
    table.print-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 9.5px !important; /* Fonte ligeiramente menor */
        color: #000000 !important;
        font-family: 'IBM Plex Sans', sans-serif;
        line-height: 1.05 !important;
        display: table !important;
        margin-bottom: 3px !important; /* Menos margem entre tabelas */
    }
    table.print-table th, table.print-table td {
        border: 1px solid #000000 !important;
        padding: 1px 3px !important; /* Espaçamento interno mais espremido */
        text-align: left;
        color: #000000 !important;
        background-color: #ffffff !important;
    }
    table.print-table th {
        background-color: #e0e0e0 !important;
        font-weight: bold;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }
    table.print-table tr { break-inside: avoid !important; page-break-inside: avoid !important; }
}
@media screen {
    #print-section { display: none !important; }
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTES E PRODUTOS INICIAIS
# ─────────────────────────────────────────────
LOJAS = ["Loja 01", "Loja 02", "Loja 03", "Loja 04", "Loja 05", "Loja 06", "Loja 07", "Loja 08"]
MAPA_LOJAS = {l: l for l in LOJAS}

produtos_iniciais = [
    {"Tipo": "Bovinos", "Descrição": "Boi Casado"},
    {"Tipo": "Bovinos", "Descrição": "Boi Dianteiro"},
    {"Tipo": "Bovinos", "Descrição": "Boi Capote"},
    {"Tipo": "Bovinos", "Descrição": "Boi Bola"},
    {"Tipo": "Bovinos", "Descrição": "Boi Bisteca Inteira"},
    {"Tipo": "Bovinos", "Descrição": "Boi Costela Inteira"},
    {"Tipo": "Bovinos", "Descrição": "Novilha Casada"},
    {"Tipo": "Bovinos", "Descrição": "Novilha Dianteiro"},
    {"Tipo": "Bovinos", "Descrição": "Novilha Capote"},
    {"Tipo": "Bovinos", "Descrição": "Novilha Bola"},
    {"Tipo": "Bovinos", "Descrição": "Novilha Bisteca Inteira"},
    {"Tipo": "Bovinos", "Descrição": "Novilha Costela Inteira"},
    
    {"Tipo": "Miúdos", "Descrição": "Bucho Kg"},
    {"Tipo": "Miúdos", "Descrição": "Coração Kg"},
    {"Tipo": "Miúdos", "Descrição": "Fígado Kg"},
    {"Tipo": "Miúdos", "Descrição": "Língua Kg"},
    {"Tipo": "Miúdos", "Descrição": "Rabada Kg"},
    {"Tipo": "Miúdos", "Descrição": "Rim Kg"},
    
    {"Tipo": "Porcos", "Descrição": "Carcaça"},
    {"Tipo": "Porcos", "Descrição": "Pé Kg"},
    {"Tipo": "Porcos", "Descrição": "Orelha Kg"},
    
    {"Tipo": "Etiquetas", "Descrição": "Bovino P"},
    {"Tipo": "Etiquetas", "Descrição": "Bovino G"},
    {"Tipo": "Etiquetas", "Descrição": "Suíno"},
    {"Tipo": "Etiquetas", "Descrição": "Avental"},
    {"Tipo": "Etiquetas", "Descrição": "Boné"},
]

# ─────────────────────────────────────────────
# CONEXÃO GOOGLE SHEETS & FUNÇÕES DE DADOS
# ─────────────────────────────────────────────
conn = st.connection("gsheets", type=GSheetsConnection)

WS_PRODUTOS = "AcPecas_Produtos"
WS_PEDIDOS  = "AcPecas"

@st.cache_data(ttl=15)
def carregar_catalogo_acpecas():
    df = conn.read(worksheet=WS_PRODUTOS, ttl=0, usecols=list(range(20)))
    
    if df.empty or "Tipo" not in df.columns:
        df_init = pd.DataFrame(produtos_iniciais)
        for loja in LOJAS: df_init[loja] = True
        conn.update(worksheet=WS_PRODUTOS, data=df_init)
        return df_init
    
    novas_colunas = {}
    for col in df.columns:
        col_str = str(col).strip()
        for loja in LOJAS:
            if loja.lower() in col_str.lower():
                novas_colunas[col] = loja
    df = df.rename(columns=novas_colunas)
    
    for loja in LOJAS:
        if loja not in df.columns:
            df[loja] = False
        else:
            def parse_bool(x):
                if isinstance(x, bool):
                    return x
                if isinstance(x, (int, float)):
                    return bool(x) and not pd.isna(x)
                return str(x).strip().upper() in ['TRUE', 'VERDADEIRO', '1', 'V', 'SIM', 'YES', 'T', 'X']
            df[loja] = df[loja].apply(parse_bool)
            
    return df

@st.cache_data(ttl=15)
def carregar_pedidos():
    df_pedidos = conn.read(worksheet=WS_PEDIDOS, ttl=0)
    df_cat = carregar_catalogo_acpecas()
    
    if not df_pedidos.empty:
        novas_colunas_ped = {}
        for col in df_pedidos.columns:
            col_str = str(col).strip()
            for loja in LOJAS:
                if loja.lower() in col_str.lower():
                    novas_colunas_ped[col] = loja
        df_pedidos = df_pedidos.rename(columns=novas_colunas_ped)
    
    if df_pedidos.empty or "Tipo" not in df_pedidos.columns:
        df_init = df_cat[["Tipo", "Descrição"]].copy()
        for loja in LOJAS:
            df_init[loja] = 0
        if not df_init.empty:
            conn.update(worksheet=WS_PEDIDOS, data=df_init)
        return df_init
        
    for loja in LOJAS:
        if loja in df_pedidos.columns:
            df_pedidos[loja] = pd.to_numeric(df_pedidos[loja], errors='coerce').fillna(0).astype(int)
            
    return df_pedidos

def salvar_pedidos(df_to_save):
    conn.update(worksheet=WS_PEDIDOS, data=df_to_save)
    st.cache_data.clear()

def salvar_catalogo(df_to_save):
    conn.update(worksheet=WS_PRODUTOS, data=df_to_save)
    st.cache_data.clear()

# ─────────────────────────────────────────────
# SISTEMA DE LOGIN
# ─────────────────────────────────────────────
if st.session_state['usuario_logado_acpecas'] is None:
    st.write("<br><br>", unsafe_allow_html=True)
    _, col2, _ = st.columns([1, 1.4, 1])
    with col2:
        with st.container(border=True):
            h1, h2 = st.columns([4, 1])
            with h1:
                st.markdown("""
                    <h2 style='margin-bottom:0'>Portal de Pedidos</h2>
                    <p style='color:#7d8590;font-size:14px;margin-top:4px'>Açougue Peças — Molicenter</p>
                """, unsafe_allow_html=True)
            with h2:
                st.write("")
                try:
                    st.image("passaro_logo.png", width=60)
                except Exception:
                    st.markdown("🥩", unsafe_allow_html=True)

            st.divider()
            usuarios_permitidos = ["Selecione..."] + ["Administrador"] + LOJAS
            usuario_selecionado = st.selectbox("👤 Usuário de acesso:", usuarios_permitidos)
            
            senha_digitada = st.text_input("🔑 Senha de acesso:", type="password", autocomplete="off")
            
            st.write("<br>", unsafe_allow_html=True)

            if st.button("Entrar no Sistema", type="primary", use_container_width=True):
                if usuario_selecionado == "Selecione...":
                    st.error("⚠️ Por favor, selecione um usuário.")
                elif usuario_selecionado == "Administrador" and senha_digitada == "moli0000":
                    st.session_state['usuario_logado_acpecas'] = usuario_selecionado
                    st.rerun()
                elif usuario_selecionado in LOJAS and senha_digitada == "moli1234":
                    st.session_state['usuario_logado_acpecas'] = usuario_selecionado
                    st.rerun()
                elif senha_digitada:
                    st.error("⚠️ Senha incorreta. Tente novamente.")

            st.markdown('<p style="font-size: 11px; color: #7d8590; text-align: center; margin-top: 10px;">🔒 Acesso restrito — Molicenter © 2026</p>', unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────
# PÓS-LOGIN
# ─────────────────────────────────────────────
usuario_atual = st.session_state['usuario_logado_acpecas']
acesso_total  = usuario_atual == "Administrador"

if not acesso_total:
    st.markdown("""
    <style>
        section[data-testid="stSidebar"] { display: none !important; }
        [data-testid="collapsedControl"]  { display: none !important; }
        .main .block-container { max-width: 100% !important; padding-left: 2.5rem !important; padding-right: 2.5rem !important; }
    </style>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    try:
        st.image("passaro_logo.png", width=72)
    except Exception:
        st.markdown("🥩")

    st.markdown(f"### Olá, *{usuario_atual}*")
    st.caption("Sistema de Pedidos — Açougue Peças")
    st.divider()

    if acesso_total:
        perfil_navegacao = st.radio("📍 Navegação:", [
            "Separação e Fechamento",
            "Visão das Lojas",
            "Visão por Tipo (Resumo)",
            "Catálogo de Produtos" 
        ])
    else:
        perfil_navegacao = "Visão das Lojas"

    st.divider()
    
    df_ped = carregar_pedidos()
    if not df_ped.empty and set(LOJAS).issubset(df_ped.columns):
        total_preenchidos = (df_ped[LOJAS] > 0).any(axis=1).sum()
    else:
        total_preenchidos = 0
        
    st.metric("Itens c/ pedido", total_preenchidos, help="Itens com ao menos 1 quantidade preenchida")
    
    st.divider()
    
    if st.button("🔄 Sincronizar Dados", use_container_width=True):
        st.cache_data.clear()
        st.session_state['reset_counter_acpecas'] += 1
        st.rerun()
        
    st.write("<br>", unsafe_allow_html=True)

    if st.button("🚪 Sair / Logout", use_container_width=True):
        st.session_state['usuario_logado_acpecas'] = None
        st.rerun()

# ─────────────────────────────────────────────
# FUNÇÃO MODAL DE CONFIRMAÇÃO PARA ZERAR
# ─────────────────────────────────────────────
@st.dialog("🚨 Confirmação Necessária")
def modal_zerar_pedidos():
    st.markdown("Tem certeza que deseja **zerar todos os pedidos** de todas as lojas?")
    st.markdown("⚠️ *Esta ação irá zerar as quantidades diretamente no Google Sheets.*")
    
    st.write("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("❌ Não, cancelar", use_container_width=True):
            st.rerun()
    with c2:
        if st.button("✔️ Sim, zerar tudo", type="primary", use_container_width=True):
            st.session_state['reset_counter_acpecas'] += 1
            df_main = carregar_pedidos()
            for loja in LOJAS:
                if loja in df_main.columns:
                    df_main[loja] = 0
            salvar_pedidos(df_main)
            st.rerun()

# ─────────────────────────────────────────────
# ROTA 1 — SEPARAÇÃO E FECHAMENTO (Admin)
# ─────────────────────────────────────────────
if perfil_navegacao == "Separação e Fechamento":
    st.markdown("""
    <div class="page-header hide-print" style="background: linear-gradient(90deg, var(--red-dark) 0%, #1a0808 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">📊</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Separação e Fechamento — Açougue Peças</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Consolidado geral de quantidades por Tipo</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.container(border=True):
        df_base = carregar_pedidos()
        
        if df_base.empty:
            st.warning("A base de pedidos está vazia. Cadastre produtos no Catálogo primeiro.")
            st.stop()
            
        df_base["TOTAL GERAL"] = df_base[LOJAS].sum(axis=1)

        col_cfg = {
            "Tipo":        st.column_config.TextColumn("Tipo", disabled=True),
            "Descrição":   st.column_config.TextColumn("Produto", disabled=True),
            "TOTAL GERAL": st.column_config.NumberColumn("TOTAL ▶️", width=90, format="%d", disabled=True),
        }
        for loja, novo_nome in MAPA_LOJAS.items():
            col_cfg[loja] = st.column_config.NumberColumn(novo_nome, format="%d", min_value=0, step=1)

        cols_order = ["Tipo", "Descrição"] + LOJAS + ["TOTAL GERAL"]
        df_exibir = df_base[cols_order]

        df_editado = st.data_editor(
            df_exibir,
            hide_index=True,
            use_container_width=True,
            height=600,
            column_config=col_cfg,
            key=f"sep_editor_{st.session_state['reset_counter_acpecas']}"
        )

        html_table = df_editado.to_html(index=False, classes="print-table")
        st.markdown(f"""<div id="print-section">
<h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
    Resumo de Separação — Açougue Peças
</h2>
<div class="print-container">
{html_table}
</div>
</div>""", unsafe_allow_html=True)

        st.divider()
        col_salvar, col_csv, col_excel, col_print, col_zerar = st.columns([2.5, 1.2, 1.2, 1.5, 2.5])

        with col_salvar:
            if st.button("💾 Salvar Alterações", type="primary", use_container_width=True):
                df_to_save = df_editado.drop(columns=["TOTAL GERAL"])
                salvar_pedidos(df_to_save)
                st.success("✅ Pedidos salvos na nuvem com sucesso!")
                st.rerun()

        with col_csv:
            df_csv = df_editado.copy().rename(columns=MAPA_LOJAS)
            csv = df_csv.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️ CSV", data=csv, file_name="separacao_acpecas.csv", mime="text/csv", use_container_width=True)

        with col_excel:
            buffer = io.BytesIO()
            
            # Prepara os dados
            df_exp = df_editado.copy().rename(columns=MAPA_LOJAS)
            
            # Cria o workbook e a planilha ativa
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedidos AcPecas"
            
            # 🎨 ESTILOS DE FORMATAÇÃO
            fill_dark_red = PatternFill(start_color="7B1315", end_color="7B1315", fill_type="solid") 
            fill_light_red = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") 
            
            font_white_bold = Font(bold=True, color="FFFFFF", size=11)
            
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            colunas = list(df_exp.columns)
            max_col = len(colunas)
            
            # 1. LINHA GERAL SUPERIOR
            hoje = datetime.datetime.now().strftime("%d/%m/%Y")
            
            # A1 e B1
            c1 = ws.cell(row=1, column=1, value="Tipo")
            c1.fill, c1.font, c1.alignment, c1.border = fill_dark_red, font_white_bold, align_center, thin_border
            
            c2 = ws.cell(row=1, column=2, value="Descrição")
            c2.fill, c2.font, c2.alignment, c2.border = fill_dark_red, font_white_bold, align_center, thin_border
            
            # C1 até final (Data)
            ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=max_col)
            c_data = ws.cell(row=1, column=3, value=f"Pedidos do dia {hoje}")
            c_data.fill, c_data.font, c_data.alignment, c_data.border = fill_dark_red, font_white_bold, align_center, thin_border
            
            # Aplicar bordas e preenchimento nas células mescladas
            for c in range(3, max_col + 1):
                cell = ws.cell(row=1, column=c)
                cell.border = thin_border
                cell.fill = fill_dark_red

            # 2. INSERINDO GRUPOS E DADOS
            row_idx = 2
            tipos = df_exp['Tipo'].unique()
            
            for tipo in tipos:
                # CABEÇALHO DO GRUPO (Ex: Pedidos Peça)
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                c_grupo = ws.cell(row=row_idx, column=1, value=f"Pedidos {tipo}")
                c_grupo.fill, c_grupo.font, c_grupo.alignment, c_grupo.border = fill_dark_red, font_white_bold, align_center, thin_border
                
                # Garantir formatação da célula B do merge
                ws.cell(row=row_idx, column=2).border = thin_border
                ws.cell(row=row_idx, column=2).fill = fill_dark_red
                
                # Nome das Lojas na mesma linha do grupo
                for col_num in range(3, max_col + 1):
                    col_name = colunas[col_num - 1]
                    c_loja = ws.cell(row=row_idx, column=col_num, value=col_name)
                    c_loja.fill, c_loja.font, c_loja.alignment, c_loja.border = fill_dark_red, font_white_bold, align_center, thin_border
                
                row_idx += 1
                
                # DADOS DO GRUPO (Zebrado)
                df_tipo = df_exp[df_exp['Tipo'] == tipo]
                linha_zebra = 0
                for _, row_data in df_tipo.iterrows():
                    linha_zebra += 1
                    for col_num in range(1, max_col + 1):
                        col_name = colunas[col_num - 1]
                        val = row_data[col_name]
                        # --- AJUSTE AQUI ---
                        # Se for coluna de quantidade (Lojas) e o valor for 0, escreve vazio
                        if (col_name in LOJAS or col_name == "TOTAL GERAL") and (val == 0 or val == 0.0):
                            cell = ws.cell(row=row_idx, column=col_num, value="")
                        else:
                            cell = ws.cell(row=row_idx, column=col_num, value=val)
                        #cell = ws.cell(row=row_idx, column=col_num, value=val)
                        cell.border = thin_border 
                        
                        # Aplica o vermelho claro (zebra) em linhas ímpares
                        if linha_zebra % 2 != 0:
                            cell.fill = fill_light_red
                        
                        # Alinhamento à esquerda para texto, centro para números
                        if col_name in ['Tipo', 'Descrição']:
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_center
                    row_idx += 1
                    
            # 4. AJUSTE DE LARGURA DAS COLUNAS
            ws.column_dimensions['A'].width = 15 # Tipo
            ws.column_dimensions['B'].width = 40 # Descrição
            for i in range(3, max_col + 1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = 12
                
            wb.save(buffer)
            
            st.download_button(
                label="⬇️ Excel",
                data=buffer.getvalue(),
                file_name="separacao_acpecas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
                               
        with col_print:
            if st.button("🖨️ Imprimir", use_container_width=True):
                components.html("<script>window.parent.print();</script>", height=0)

        with col_zerar:
            if st.button("🚨 Zerar Todos os Pedidos", use_container_width=True):
                modal_zerar_pedidos()

# ─────────────────────────────────────────────
# ROTA 2 — VISÃO DAS LOJAS
# ─────────────────────────────────────────────
elif perfil_navegacao == "Visão das Lojas":
    if acesso_total:
        loja_selecionada = st.selectbox("👁️ Visualizar como:", LOJAS)
    else:
        loja_selecionada = usuario_atual

    col_info, col_logout = st.columns([8, 2])
    with col_info:
        id_loja = MAPA_LOJAS.get(loja_selecionada, loja_selecionada)
        st.markdown(f"""
        <div class="topbar-loja hide-print">
            <div class="topbar-left">
                <span style="font-size:22px">🥩</span>
                <div>
                    <div class="topbar-title">{loja_selecionada} — Açougue Peças</div>
                    <div class="topbar-sub">Preencha a quantidade de cada produto</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col_logout:
        st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
        if st.button("🚪 Sair / Logout", use_container_width=True):
            st.session_state['usuario_logado_acpecas'] = None
            st.rerun()

    df_cat = carregar_catalogo_acpecas()
    df_cat_loja = df_cat[df_cat[loja_selecionada] == True].copy()
    
    if df_cat_loja.empty:
        st.warning(f"Nenhum produto habilitado para a {loja_selecionada} no momento.")
        st.stop()

    df_all = carregar_pedidos()
    df_loja_view = pd.merge(
        df_cat_loja[["Tipo", "Descrição"]],
        df_all[["Tipo", "Descrição", loja_selecionada]],
        on=["Tipo", "Descrição"],
        how="left"
    )
    df_loja_view[loja_selecionada] = df_loja_view[loja_selecionada].fillna(0).astype(int)
    df_loja_view = df_loja_view.rename(columns={loja_selecionada: "Qtde"})

    with st.container(border=True):
        st.info("💡 Preencha a *Qtde* desejada para cada produto. Apenas os itens atrelados a esta loja são exibidos.")

        col_cfg_loja = {
            "Tipo":       st.column_config.TextColumn("Tipo", width=150, disabled=True),
            "Descrição":  st.column_config.TextColumn("Produto", width=250, disabled=True),
            "Qtde":       st.column_config.NumberColumn("🛒 Qtde", width=120, min_value=0, step=1),
        }

        with st.container():
            df_editado = st.data_editor(
                df_loja_view,
                column_config=col_cfg_loja,
                hide_index=True,
                use_container_width=True,
                height=520,
                key=f"loja_acpecas_{st.session_state['reset_counter_acpecas']}"
            )

        # ── HTML INVISÍVEL PARA IMPRESSÃO DA LOJA ─────────────────────
        html_table_loja = df_editado.to_html(index=False, classes="print-table")
        st.markdown(f"""<div id="print-section">
<h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
    Resumo do Pedido — {loja_selecionada}
</h2>
<div class="print-container">
{html_table_loja}
</div>
</div>""", unsafe_allow_html=True)
        # ──────────────────────────────────────────────────────────────

        itens_com_pedido = int((df_editado["Qtde"] > 0).sum())
        total_itens      = len(df_editado)
        total_unidades   = int(df_editado["Qtde"].sum())
        pct              = round(itens_com_pedido / total_itens * 100) if total_itens else 0

        st.divider()
        m1, m2, m3, col_print, col_btn = st.columns([2.5, 2.2, 1.8, 1.5, 3])
        with m1: st.metric("Itens preenchidos", f"{itens_com_pedido} / {total_itens}")
        with m2: st.metric("Total de unidades", total_unidades)
        with m3: st.metric("Cobertura", f"{pct}%")
        
        with col_print:
            st.write("<br>", unsafe_allow_html=True)
            if st.button("🖨️ Imprimir", use_container_width=True):
                components.html("<script>window.parent.print();</script>", height=0)

        with col_btn:
            st.write("<br>", unsafe_allow_html=True)
            if st.button("💾 Salvar Pedido da Semana", type="primary", use_container_width=True):
                df_main = carregar_pedidos()
                
                for _, row in df_editado.iterrows():
                    mask = (
                        (df_main["Tipo"] == row["Tipo"]) &
                        (df_main["Descrição"] == row["Descrição"])
                    )
                    if mask.any():
                        df_main.loc[mask, loja_selecionada] = row["Qtde"]
                    else:
                        nova_linha = {"Tipo": row["Tipo"], "Descrição": row["Descrição"]}
                        for l in LOJAS: nova_linha[l] = 0
                        nova_linha[loja_selecionada] = row["Qtde"]
                        df_main = pd.concat([df_main, pd.DataFrame([nova_linha])], ignore_index=True)
                
                salvar_pedidos(df_main)
                st.success(f"✅ Pedido da {loja_selecionada} enviado para a nuvem com sucesso!")

# ─────────────────────────────────────────────
# ROTA 3 — VISÃO POR TIPO / RESUMO (Admin)
# ─────────────────────────────────────────────
elif perfil_navegacao == "Visão por Tipo (Resumo)":
    st.markdown("""
    <div class="hide-print" style="background: linear-gradient(90deg, var(--red-dark) 0%, #1a0808 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">🥩</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Visão por Tipo — Açougue Peças</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Resumo consolidado agrupado pelas categorias de produtos</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    df_all = carregar_pedidos()
    df_cat = carregar_catalogo_acpecas()
    
    if df_all.empty or df_cat.empty:
        st.warning("Não há dados de pedidos ou catálogo preenchidos.")
        st.stop()

    nomes_tipos = df_cat["Tipo"].dropna().unique().tolist()
    
    # Variável para acumular o HTML de todas as tabelas para impressão
    html_print_content = ""

    for i in range(0, len(nomes_tipos), 1):
        cols = st.columns(1, gap="small")
        for j, tipo_prod in enumerate(nomes_tipos[i:i+1]):

            # Carrega a base geral para o Tipo atual
            df_forn = df_all[df_all["Tipo"] == tipo_prod].copy()
            
            # Forçamos mostrar sempre as colunas
            colunas_presentes = LOJAS
            df_forn = df_forn[["Descrição"] + colunas_presentes].copy()
            df_forn["TOTAL"] = df_forn[colunas_presentes].sum(axis=1)
            
            lojas_renomeadas = {l: MAPA_LOJAS[l] for l in colunas_presentes}
            df_forn = df_forn.rename(columns=lojas_renomeadas)

            lojas_cols_renomeadas = [MAPA_LOJAS[l] for l in colunas_presentes]

            col_cfg_forn = {
                "Descrição": st.column_config.TextColumn("Produto", disabled=False),
                "TOTAL":     st.column_config.NumberColumn("TOTAL", format="%d", disabled=True),
            }
            for c in lojas_cols_renomeadas:
                col_cfg_forn[c] = st.column_config.NumberColumn(c, format="%d", disabled=False, min_value=0)

            altura = (len(df_forn) * 35) + 42 

            with cols[j]:
                with st.container(border=True):
                    st.markdown('<div class="title-input">', unsafe_allow_html=True)
                    st.text_input(
                        "Tipo",
                        value=f"🥩 {tipo_prod}",
                        label_visibility="collapsed",
                        key=f"title_tipo_{tipo_prod}_{st.session_state['reset_counter_acpecas']}"
                    )
                    st.markdown('</div>', unsafe_allow_html=True)

                    cols_order_forn = ["Descrição"] + lojas_cols_renomeadas + ["TOTAL"]
                    df_forn_edit = st.data_editor(
                        df_forn[cols_order_forn],
                        hide_index=True,
                        use_container_width=True,
                        column_config=col_cfg_forn,
                        height=altura,
                        num_rows="fixed",
                        key=f"tipo_acpecas_{tipo_prod}_{st.session_state['reset_counter_acpecas']}"
                    )

                    total_geral = int(df_forn_edit["TOTAL"].sum()) if "TOTAL" in df_forn_edit.columns else 0
                    st.markdown(f"""
                        <div style="text-align:right; font-weight:700; margin-top:6px; color:var(--red-bright); font-size:15px;">
                            Total Geral: {total_geral} unidades
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Acumulando HTML para a página de impressão SEM ESPAÇOS INICIAIS
                    html_table = df_forn_edit.to_html(index=False, classes="print-table")
                    html_print_content += f"<h3 style='color: black; margin-top: 10px; margin-bottom: 4px;'>🥩 {tipo_prod}</h3>\n"
                    html_print_content += f"{html_table}\n"
                    html_print_content += f"<div style='text-align:right; font-weight:bold; font-size:11px; margin-top:3px; margin-bottom: 8px; color: black;'>Total da Categoria: {total_geral} unidades</div>\n"

        st.write("<br>", unsafe_allow_html=True)

    # ── HTML INVISÍVEL PARA IMPRESSÃO DO RESUMO ───────────────────
    st.markdown(f"""<div id="print-section">
<h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
    Visão por Tipo (Resumo) — Açougue Peças
</h2>
<div class="print-container">
{html_print_content}
</div>
</div>""", unsafe_allow_html=True)
    # ──────────────────────────────────────────────────────────────
    
    st.divider()
    _, col_print = st.columns([8, 2])
    with col_print:
        if st.button("🖨️ Imprimir Resumo Geral", use_container_width=True):
            components.html("<script>window.parent.print();</script>", height=0)

# ─────────────────────────────────────────────
# ROTA 4 — CATÁLOGO DE PRODUTOS
# ─────────────────────────────────────────────
elif perfil_navegacao == "Catálogo de Produtos":
    st.markdown("""
    <div class="hide-print" style="background: linear-gradient(90deg, var(--red-dark) 0%, #1a0808 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">🗂️</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Catálogo de Peças</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Gerencie itens e controle a visibilidade por loja através das caixas de seleção</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    df_catalogo = carregar_catalogo_acpecas()
    
    with st.container(border=True):
        
        col_cfg_cat = {
            "Tipo":       st.column_config.SelectboxColumn("Tipo (Categoria)", options=["Bovinos", "Miúdos", "Porcos", "Etiquetas"], width=150, required=True),
            "Descrição":  st.column_config.TextColumn("Descrição do Item", width=350, required=True),
        }
        for loja in LOJAS:
            col_cfg_cat[loja] = st.column_config.CheckboxColumn(loja, default=False)
            
        edited_cat = st.data_editor(
            df_catalogo,
            use_container_width=True,
            hide_index=True,
            height=600,
            num_rows="dynamic",
            column_config=col_cfg_cat,
            key=f"editor_catalogo_acpecas_{st.session_state['reset_counter_acpecas']}"
        )
        
        st.divider()
        if st.button("💾 Salvar Catálogo no Google Sheets", type="primary", use_container_width=True):
            salvar_catalogo(edited_cat)
            st.session_state['reset_counter_acpecas'] += 1
            st.success("✅ Catálogo atualizado com sucesso! As alterações já refletem na visão das lojas.")
            st.rerun()
